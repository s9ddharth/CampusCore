from __future__ import annotations

from io import BytesIO
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from fastapi import HTTPException, status
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import (
    SemesterResult,
    Student,
    StudentResult,
    Subject,
)


# =========================================================
# HELPERS
# =========================================================

def get_student(
    db: Session,
    student_id: int,
) -> Student:

    student = db.scalar(
        select(Student).where(
            Student.id == student_id
        )
    )

    if student is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student not found.",
        )

    return student


def get_student_results(
    db: Session,
    *,
    student_id: int,
    semester: int | None = None,
    academic_year: str | None = None,
) -> list[StudentResult]:

    conditions = [
        StudentResult.student_id == student_id
    ]

    if semester is not None:
        conditions.append(
            StudentResult.semester == semester
        )

    if academic_year is not None:
        conditions.append(
            StudentResult.academic_year
            == academic_year
        )

    return list(
        db.scalars(
            select(StudentResult)
            .where(*conditions)
            .order_by(
                StudentResult.semester.asc(),
                StudentResult.subject_id.asc(),
            )
        ).all()
    )


def get_subject_map(
    db: Session,
    results: list[StudentResult],
) -> dict[int, Subject]:

    subject_ids = {
        result.subject_id
        for result in results
    }

    if not subject_ids:
        return {}

    subjects = db.scalars(
        select(Subject).where(
            Subject.id.in_(subject_ids)
        )
    ).all()

    return {
        subject.id: subject
        for subject in subjects
    }


def get_gpa_map(
    db: Session,
    student_id: int,
) -> dict[tuple[int, str], SemesterResult]:

    semester_results = db.scalars(
        select(SemesterResult).where(
            SemesterResult.student_id == student_id
        )
    ).all()

    return {
        (
            result.semester,
            result.academic_year,
        ): result
        for result in semester_results
    }


# =========================================================
# EXCEL REPORT
# =========================================================

def generate_student_result_excel(
    db: Session,
    *,
    student_id: int,
    semester: int | None = None,
    academic_year: str | None = None,
) -> BytesIO:

    student = get_student(
        db,
        student_id,
    )

    results = get_student_results(
        db,
        student_id=student_id,
        semester=semester,
        academic_year=academic_year,
    )

    if not results:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No result records found.",
        )

    subjects = get_subject_map(
        db,
        results,
    )

    gpa_map = get_gpa_map(
        db,
        student_id,
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Academic Results"

    # -----------------------------------------------------
    # Header
    # -----------------------------------------------------

    worksheet["A1"] = "CampusCore"
    worksheet["A1"].font = Font(
        bold=True,
        size=18,
    )

    worksheet["A2"] = (
        "Student Academic Result Report"
    )
    worksheet["A2"].font = Font(
        bold=True,
        size=14,
    )

    worksheet["A4"] = "Student ID"
    worksheet["B4"] = student.id

    worksheet["A5"] = "Roll No"
    worksheet["B5"] = student.roll_no

    worksheet["A6"] = "Name"
    worksheet["B6"] = student.name

    worksheet["A7"] = "Semester"
    worksheet["B7"] = (
        semester
        if semester is not None
        else "All"
    )

    worksheet["A8"] = "Academic Year"
    worksheet["B8"] = (
        academic_year
        if academic_year is not None
        else "All"
    )

    # -----------------------------------------------------
    # Table
    # -----------------------------------------------------

    start_row = 10

    headers = [
        "Semester",
        "Academic Year",
        "Subject Code",
        "Subject Name",
        "Credits",
        "Raw Total",
        "TEE",
        "Normalized Score",
        "Grade",
        "Grade Point",
        "Rank",
        "Status",
    ]

    for column, header in enumerate(
        headers,
        start=1,
    ):
        cell = worksheet.cell(
            row=start_row,
            column=column,
            value=header,
        )

        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center"
        )

    row = start_row + 1

    for result in results:
        subject = subjects.get(
            result.subject_id
        )

        worksheet.cell(
            row=row,
            column=1,
            value=result.semester,
        )

        worksheet.cell(
            row=row,
            column=2,
            value=result.academic_year,
        )

        worksheet.cell(
            row=row,
            column=3,
            value=(
                subject.code
                if subject
                else result.subject_id
            ),
        )

        worksheet.cell(
            row=row,
            column=4,
            value=(
                subject.name
                if subject
                else "Unknown Subject"
            ),
        )

        worksheet.cell(
            row=row,
            column=5,
            value=(
                float(subject.credits)
                if subject
                else 0
            ),
        )

        worksheet.cell(
            row=row,
            column=6,
            value=float(
                result.raw_total
            ),
        )

        worksheet.cell(
            row=row,
            column=7,
            value=float(
                result.tee_score
            ),
        )

        worksheet.cell(
            row=row,
            column=8,
            value=float(
                result.normalized_score
            ),
        )

        worksheet.cell(
            row=row,
            column=9,
            value=result.grade,
        )

        worksheet.cell(
            row=row,
            column=10,
            value=float(
                result.grade_point
            ),
        )

        worksheet.cell(
            row=row,
            column=11,
            value=result.rank,
        )

        worksheet.cell(
            row=row,
            column=12,
            value=result.status,
        )

        row += 1

    # -----------------------------------------------------
    # GPA summary
    # -----------------------------------------------------

    row += 2

    worksheet.cell(
        row=row,
        column=1,
        value="Semester GPA Summary",
    ).font = Font(
        bold=True,
        size=13,
    )

    row += 1

    summary_headers = [
        "Semester",
        "Academic Year",
        "GPA",
        "Total Credits",
        "Status",
    ]

    for column, header in enumerate(
        summary_headers,
        start=1,
    ):
        cell = worksheet.cell(
            row=row,
            column=column,
            value=header,
        )

        cell.font = Font(bold=True)

    row += 1

    summary_items = sorted(
        gpa_map.values(),
        key=lambda item: (
            item.semester,
            item.academic_year,
        ),
    )

    for semester_result in summary_items:

        if semester is not None:
            if (
                semester_result.semester
                != semester
            ):
                continue

        if academic_year is not None:
            if (
                semester_result.academic_year
                != academic_year
            ):
                continue

        worksheet.cell(
            row=row,
            column=1,
            value=semester_result.semester,
        )

        worksheet.cell(
            row=row,
            column=2,
            value=semester_result.academic_year,
        )

        worksheet.cell(
            row=row,
            column=3,
            value=float(
                semester_result.gpa
            ),
        )

        worksheet.cell(
            row=row,
            column=4,
            value=float(
                semester_result.total_credits
            ),
        )

        worksheet.cell(
            row=row,
            column=5,
            value=semester_result.status,
        )

        row += 1

    # -----------------------------------------------------
    # Formatting
    # -----------------------------------------------------

    widths = [
        12,
        16,
        14,
        28,
        10,
        12,
        10,
        18,
        10,
        12,
        8,
        14,
    ]

    for index, width in enumerate(
        widths,
        start=1,
    ):
        worksheet.column_dimensions[
            get_column_letter(index)
        ].width = width

    worksheet.freeze_panes = "A11"

    output = BytesIO()

    workbook.save(output)
    output.seek(0)

    return output


# =========================================================
# PDF REPORT
# =========================================================

def generate_student_result_pdf(
    db: Session,
    *,
    student_id: int,
    semester: int | None = None,
    academic_year: str | None = None,
) -> BytesIO:

    student = get_student(
        db,
        student_id,
    )

    results = get_student_results(
        db,
        student_id=student_id,
        semester=semester,
        academic_year=academic_year,
    )

    if not results:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No result records found.",
        )

    subjects = get_subject_map(
        db,
        results,
    )

    gpa_map = get_gpa_map(
        db,
        student_id,
    )

    output = BytesIO()

    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=25,
        leftMargin=25,
        topMargin=25,
        bottomMargin=25,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "TitleStyle",
        parent=styles["Title"],
        alignment=TA_CENTER,
        fontSize=18,
        spaceAfter=8,
    )

    subtitle_style = ParagraphStyle(
        "SubtitleStyle",
        parent=styles["Normal"],
        alignment=TA_CENTER,
        fontSize=12,
        spaceAfter=12,
    )

    story = []

    story.append(
        Paragraph(
            "CampusCore",
            title_style,
        )
    )

    story.append(
        Paragraph(
            "Student Academic Result Report",
            subtitle_style,
        )
    )

    student_info = [
        [
            "Student ID",
            str(student.id),
            "Roll No",
            student.roll_no,
        ],
        [
            "Name",
            student.name,
            "Semester",
            str(semester)
            if semester is not None
            else "All",
        ],
        [
            "Academic Year",
            academic_year
            if academic_year is not None
            else "All",
            "Status",
            student.status,
        ],
    ]

    info_table = Table(
        student_info,
        colWidths=[
            80,
            200,
            90,
            180,
        ],
    )

    info_table.setStyle(
        TableStyle(
            [
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, -1),
                    "Helvetica",
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (0, -1),
                    "Helvetica-Bold",
                ),
                (
                    "FONTNAME",
                    (2, 0),
                    (2, -1),
                    "Helvetica-Bold",
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey,
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE",
                ),
                (
                    "BACKGROUND",
                    (0, 0),
                    (0, -1),
                    colors.whitesmoke,
                ),
                (
                    "BACKGROUND",
                    (2, 0),
                    (2, -1),
                    colors.whitesmoke,
                ),
            ]
        )
    )

    story.append(info_table)
    story.append(Spacer(1, 15))

    # -----------------------------------------------------
    # Results table
    # -----------------------------------------------------

    data = [
        [
            "Sem",
            "Year",
            "Subject",
            "Credits",
            "Raw",
            "TEE",
            "Normalized",
            "Grade",
            "Point",
            "Rank",
            "Status",
        ]
    ]

    for result in results:
        subject = subjects.get(
            result.subject_id
        )

        data.append(
            [
                str(result.semester),
                result.academic_year,
                (
                    subject.code
                    if subject
                    else str(result.subject_id)
                ),
                (
                    str(subject.credits)
                    if subject
                    else "0"
                ),
                str(result.raw_total),
                str(result.tee_score),
                str(result.normalized_score),
                result.grade,
                str(result.grade_point),
                (
                    str(result.rank)
                    if result.rank is not None
                    else "-"
                ),
                result.status,
            ]
        )

    result_table = Table(
        data,
        repeatRows=1,
    )

    result_table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.lightgrey,
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold",
                ),
                (
                    "ALIGN",
                    (0, 0),
                    (-1, -1),
                    "CENTER",
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey,
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE",
                ),
                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, -1),
                    8,
                ),
            ]
        )
    )

    story.append(result_table)
    story.append(Spacer(1, 15))

    # -----------------------------------------------------
    # GPA summary
    # -----------------------------------------------------

    story.append(
        Paragraph(
            "Semester GPA Summary",
            styles["Heading2"],
        )
    )

    gpa_data = [
        [
            "Semester",
            "Academic Year",
            "GPA",
            "Credits",
            "Status",
        ]
    ]

    for semester_result in sorted(
        gpa_map.values(),
        key=lambda item: (
            item.semester,
            item.academic_year,
        ),
    ):

        if semester is not None:
            if semester_result.semester != semester:
                continue

        if academic_year is not None:
            if (
                semester_result.academic_year
                != academic_year
            ):
                continue

        gpa_data.append(
            [
                str(semester_result.semester),
                semester_result.academic_year,
                str(semester_result.gpa),
                str(semester_result.total_credits),
                semester_result.status,
            ]
        )

    gpa_table = Table(
        gpa_data,
        repeatRows=1,
        colWidths=[
            70,
            120,
            70,
            80,
            100,
        ],
    )

    gpa_table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.lightgrey,
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold",
                ),
                (
                    "ALIGN",
                    (0, 0),
                    (-1, -1),
                    "CENTER",
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey,
                ),
            ]
        )
    )

    story.append(gpa_table)

    document.build(story)

    output.seek(0)

    return output